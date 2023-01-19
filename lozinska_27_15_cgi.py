import cgi
import openpyxl

page = """<html>
    <head>
        <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
	</head>

    <title>Exchange rate</title>

    <body>

        <h3>Checking exchange rate</h3>
        {}
        <p>
            <b>Choose the amonut of money you want to transfer into a different currency (avialble currencies UAH, USD, EUR, PLN): </b>
        <br>

            Monetary amount
            <input type=text name=summ value="">
            in currency
            <input type=text name=s_cur value="">
            to currency
            <input type=text name=e_cur value="">

            <input type=submit value="?">

        </p>
    </body>
</html>
"""

file = openpyxl.load_workbook("currencies.xlsx")
data = file.active

def counter(summ1, summ2):
    return summ1*summ2

form= cgi.FieldStorage()
summ = float(form["summ"].value)
s_cur = str(form["s_cur"].value)
e_cur = str(form["e_cur"].value)


for row in data.iter_rows():
    for cell in row:

        if cell.value == s_cur:
            if data.cell(row = cell.row, column = cell.column+1).value == e_cur:
                exchange = data.cell(row = cell.row, column = cell.column+2).value

                result = "{s1}{c1} = {s2}{c2}".format(s1 = summ, c1= s_cur, s2 = counter(summ, exchange), c2 = e_cur)
                print(page.format(result))