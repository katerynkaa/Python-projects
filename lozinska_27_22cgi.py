import openpyxl
import cgi
import datetime

page = """<html>
 <head>
     <meta charset="utf-8">
     <title>Сервіс пошуку рейсів</title>
 </head>
 <body>
 <h1>Оберіть аеропорти</h1>
 <form method=POST action="">
     <select name="airports1" multiple="multiple" size="4">
         <option value="Бориспіль">KBP</option>
         <option value="Київ">IEV</option>
         <option value="Шарль-де-Голль">CDG</option>
         <option value="Берлінський Інтернаціональний Аеропорт">BER</option>
     </select>
    <br>
    до
    <br>

      <select name="airports2" multiple="multiple" size="4">
          <option value="Бориспіль">KBP</option>
         <option value="Київ">IEV</option>
         <option value="Шарль-де-Голль">CDG</option>
         <option value="Берлінський Інтернаціональний Аеропорт">BER</option>
      </select>
     <br>
       на дату:<input type = text name=date value="">
     <p>
       <input type="submit" value="Знайти">
     </p>
    <br>
     <h2>Cписок рейсів:</h2>>
     <table style="width:50%">
        {}

    </table>
 </form>
 </body>
</html>"""


class Flight:

    @staticmethod
    def cutId(flight_name: str):
        for i in range(len(flight_name.lower())):
            if not flight_name.lower()[i].isalpha():
                return flight_name[:i]

    def __init__(self, Id1: str, Id2: str, planeName: str, days: str, depart_Time: str,
                 arrive_Time: str, p_Class: str, price: str):
        self.Id1 = Id1
        self.Id2 = Id2
        self.planeName: str = planeName
        self.airport_id = self.cutId(planeName)
        self.depart_Time = depart_Time
        self.arrive_Time = arrive_Time
        self.p_Class = p_Class
        self.price = float(price)
        self.days = days


class Airflights:
    def __init__(self, fileName):
        self.wb = openpyxl.load_workbook(fileName)
        self.aircompanies = {}
        self.airports = {}
        self.flights = {}

    def companyInformation(self):

        ws = self.wb['Авіакомпанії']

        for k in range(ws.min_row + 1, ws.max_row + 1):
            c = ws.cell(row=k, column=ws.min_column)
            n = ws.cell(row=k, column=ws.min_column + 1)
            self.aircompanies[c.value] = n.value

    def flightsInformation(self):

        ws = self.wb["Рейси"]

        for k in range(ws.min_row + 1, ws.max_row + 1):
            Id1 = ws.cell(row=k, column=ws.min_column).value
            Id2 = ws.cell(row=k, column=ws.min_column + 1).value
            p_Name = ws.cell(row=k, column=ws.min_column + 2).value
            days = ws.cell(row=k, column=ws.min_column + 3).value
            departTime = ws.cell(row=k, column=ws.min_column + 4).value
            arriveTime = ws.cell(row=k, column=ws.min_column + 5).value
            p_Class = ws.cell(row=k, column=ws.min_column + 6).value
            price = ws.cell(row=k, column=ws.min_column + 7).value

            self.flights[p_Name] = \
                Flight(Id1, Id2, p_Name, days, departTime,
                       arriveTime, p_Class, price)

    def airportsInformation(self):

        ws = self.wb["Аеропроти"]

        for k in range(ws.min_row + 1, ws.max_row + 1):
            code = ws.cell(row=k, column=ws.min_column)
            name = ws.cell(row=k, column=ws.min_column + 1)
            city = ws.cell(row=k, column=ws.min_column + 2)
            self.airports[code.value] = (name.value, city.value)

    def printFlight(self, flight, res, i):

        res +=   f"<tr>  " \
                 f"<th>ІНФОРМАЦІЯ</th>    " \
                 f"<th>Рейс {i}</th>    " \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Рейс:</td>" \
                 f"<td>{flight.planeName}</td>" \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Звідки -> куди:</td>" \
                 f"<td>{self.airports[flight.Id1][1]} -> {self.airports[flight.Id2][1]}</td>" \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Аеропорт</td>" \
                 f"<td>{self.airports[flight.Id1][0]} -> {self.airports[flight.Id2][0]}</td>" \
                 f"</tr>"\
                 f"<tr>" \
                 f"<td>Компанія</td>" \
                 f"<td>{self.aircompanies[flight.airport_id]}</td>" \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Відліт - приліт</td>" \
                 f"<td>{flight.depart_Time} - {flight.arrive_Time}</td>" \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Клас</td>" \
                 f"<td>{flight.p_Class}</td>" \
                 f"</tr>" \
                 f"<tr>" \
                 f"<td>Вартість</td>" \
                 f"<td>{flight.price}</td>" \
                 f"</tr>"
        return res

    def common(self, fromCity, toCity):
        allflights = []

        for key, value in self.flights.items():
            fr = self.airports[value.Id1][0]
            to = self.airports[value.Id2][0]
            if fromCity == fr and toCity == to:
                allflights.append(value)
        return allflights

    @staticmethod
    def find_flights_on_date(flight_lst):
        flights_need = []
        date = str(form['date'].value)
        c = date.split(".")
        r = datetime.datetime(int(c[2]), int(c[1]), int(c[0])).isoweekday()
        for flight in flight_lst:
            dates = str(flight.days)
            for i in dates:
                if str(r) == i:
                    flights_need.append(flight)
        return flights_need

form = cgi.FieldStorage()

table = Airflights("flights.xlsx")
table.companyInformation()
table.airportsInformation()
table.flightsInformation()

air1 = str(form.getfirst('airports1', ''))
air2 = str(form.getfirst('airports2', ''))

res = ''
while True:
    flight = table.common(air1, air2)
    if flight:
        d_flight = table.find_flights_on_date(flight)
        if d_flight:
            print('Found flights on date')
            for i in range(len(d_flight)):
                a = ''
                res += table.printFlight(d_flight[i], a, i+1)
                break
            else:
                res = f"<tr>  " \
                    f"<th>ІНФОРМАЦІЯ</th>    " \
                    f"<th>Рейс ?</th>    " \
                    f"</tr>" \
                    f"<tr>" \
                    f"<td>Рейс:</td>" \
                    f"<td>Рейси з {air1} до {air2} на вказану дату відсутні</td>" \
                    f"</tr>"
                break
        else:
            res = f"<tr>  " \
                    f"<th>ІНФОРМАЦІЯ</th>    " \
                    f"<th>Рейс ?</th>    " \
                    f"</tr>" \
                    f"<tr>" \
                    f"<td>Рейс:</td>" \
                    f"<td>Рейси з {air1} до {air2} відсутні </td>" \
                    f"</tr>"
            break

body = page.format(res)
