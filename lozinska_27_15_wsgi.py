import cgi
from wsgiref.simple_server import make_server
import openpyxl

page = """<html>
    <head>
        <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
	</head>

    <title>Exchange rate</title>

    <body>

        <h3>Checking exchange rate</h3>
        {}
        <form method =POST action=""> 
        <p>
            <b>Choose the amonut of money you want to transfer into a different currency (avialble currencies UAH, USD, EUR, PLN): </b>
        <br>

            Monetary amount
            <input type=text name=summ value="">
            in currency
            <input type=text name=s_cur value="">
            to currency
            <input type=text name=e_cur value="">

            <input type=submit value="?">

        </p>
        </form>
    </body>
</html>"""


def application(environ, start_response):

    if environ.get('PATH_INFO', '').lstrip('/'):
        result = ''
        form = cgi.FieldStorage(fp=environ['wsgi.input'],
                                environ=environ)

        summ = float(form["summ"].value)
        s_cur = str(form["s_cur"].value)
        e_cur = str(form["e_cur"].value)

        file = openpyxl.load_workbook("currencies.xlsx")
        data = file.active

        for row in data.iter_rows():
            for cell in row:

                if cell.value == s_cur:
                    if data.cell(row=cell.row, column=cell.column + 1).value == e_cur:
                        exchange = data.cell(row=cell.row, column=cell.column + 2).value

                        result += "{s1}{c1} = {s2}{c2}".format(s1=summ, c1=s_cur, s2=summ * exchange, c2=e_cur)
        body = page.format(result)
        start_response('200 OK', [('Content-Type', 'text/html; charset=utf-8')])
    else:
        start_response('404 NOT FOUND', [('Content-Type', 'text/plain')])
        body = 'Page not found'

    return [bytes(body, encoding='utf-8')]


if __name__ == '__main__':

    httpd = make_server('localhost', 1024, application)
    httpd.serve_forever()